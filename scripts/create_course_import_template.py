"""Create an Excel template for importing a course + tees + holes + yardages."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

TEMPLATE_PATH = Path("templates/course_import_template.xlsx")


def style_header(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"


def build_course_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "course"
    headers = ["course_name", "location", "notes"]
    style_header(ws, headers)
    ws.append(["Pine Valley Golf Club", "Pine Valley, NJ", "Private, very tough"])


def build_tees_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("tees")
    headers = ["tee_name", "course_rating", "slope_rating", "yardage_total"]
    style_header(ws, headers)
    ws.append(["Blue", 71.2, 128, 6900])
    ws.append(["White", 69.5, 122, 6400])
    ws.append(["Red", 67.8, 115, 5900])


def build_holes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("holes")
    headers = ["hole_number", "par"]
    style_header(ws, headers)
    for hole in range(1, 19):
        ws.append([hole, 4])


def build_tee_holes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("tee_holes")
    headers = ["tee_name", "hole_number", "yardage"]
    style_header(ws, headers)
    ws.append(["Blue", 1, 410])
    ws.append(["Blue", 2, 390])
    ws.append(["White", 1, 380])
    ws.append(["White", 2, 365])


def main() -> None:
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_course_sheet(wb)
    build_tees_sheet(wb)
    build_holes_sheet(wb)
    build_tee_holes_sheet(wb)
    wb.save(TEMPLATE_PATH)
    print(f"Template created at {TEMPLATE_PATH}")


if __name__ == "__main__":
    main()
