"""Create a clean, user-friendly Excel template for round tracking."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_PATH = Path("templates/golf_stats_template.xlsx")


def style_header(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).coordinate}"


def set_col_widths(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width


def add_validation(
    ws,
    col_letter: str,
    allowed: list[str],
    list_ws,
    list_col_counter: dict[str, int],
) -> None:
    # Excel has a strict length limit for inline list validation formulas.
    # For short lists we keep inline; for longer lists we reference a hidden sheet range.
    inline_formula = '"' + ",".join(allowed) + '"'
    if len(inline_formula) <= 240:
        formula = inline_formula
    else:
        list_col_idx = list_col_counter["value"]
        list_col_counter["value"] += 1
        list_col = get_column_letter(list_col_idx)
        start_row = 1
        for row_idx, value in enumerate(allowed, start=start_row):
            list_ws.cell(row=row_idx, column=list_col_idx, value=value)
        end_row = start_row + len(allowed) - 1
        formula = f"=_lists!${list_col}${start_row}:${list_col}${end_row}"

    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}500")


def build_rounds_sheet(wb: Workbook, list_ws, list_col_counter: dict[str, int]) -> None:
    ws = wb.active
    ws.title = "rounds"

    headers = [
        "round_external_id",
        "date_played",
        "course_name",
        "tee_name",
        "holes_played",
        "conditions",
        "round_type",
        "round_format",
        "notes",
    ]

    style_header(ws, headers)
    set_col_widths(
        ws,
        {
            1: 18,
            2: 14,
            3: 24,
            4: 14,
            5: 14,
            6: 18,
            7: 14,
            8: 14,
            9: 28,
        },
    )

    add_validation(ws, "E", ["Front 9", "Back 9", "18"], list_ws, list_col_counter)
    add_validation(ws, "G", ["Practice", "Tournament", "Casual"], list_ws, list_col_counter)
    add_validation(ws, "H", ["Stroke", "Match", "Scramble", "Other"], list_ws, list_col_counter)

    # Example row (minimal guidance)
    ws.append(
        [
            "2026-02-08-PineValley",
            "2026-02-08",
            "Pine Valley Golf Club",
            "Blue",
            "18",
            "Sunny, light wind",
            "Practice",
            "Stroke",
            "Felt good off the tee",
        ]
    )


def build_hole_stats_sheet(wb: Workbook, list_ws, list_col_counter: dict[str, int]) -> None:
    ws = wb.create_sheet("hole_stats")

    headers = [
        "round_external_id",
        "hole_number",
        "strokes",
        "putts",
        "tee_shot",
        "approach",
        "tee_club",
        "approach_club",
        "bunker_found",
        "out_of_bounds_count",
    ]

    style_header(ws, headers)
    set_col_widths(
        ws,
        {
            1: 22,
            2: 12,
            3: 10,
            4: 10,
            5: 12,
            6: 12,
            7: 12,
            8: 14,
            9: 14,
        },
    )

    add_validation(ws, "B", [str(i) for i in range(1, 19)], list_ws, list_col_counter)
    add_validation(
        ws,
        "E",
        [
            "Fairway",
            "Left",
            "Right",
            "Short",
            "Long",
            "Out Left",
            "Out Right",
            "Out Short",
            "Out Long",
            "Bunker Left",
            "Bunker Right",
            "Bunker Short",
            "Bunker Long",
            "Green",
        ],
        list_ws,
        list_col_counter,
    )
    add_validation(
        ws,
        "F",
        [
            "Green",
            "Left",
            "Right",
            "Short",
            "Long",
            "Out Left",
            "Out Right",
            "Out Short",
            "Out Long",
            "Bunker Left",
            "Bunker Right",
            "Bunker Short",
            "Bunker Long",
            "N/A",
        ],
        list_ws,
        list_col_counter,
    )

    ws.append(
        [
            "2026-02-08-PineValley",
            1,
            4,
            2,
            "Fairway",
            "Green",
            "Driver",
            "7i",
            0,
            0,
        ]
    )


def main() -> None:
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    list_ws = wb.create_sheet("_lists")
    list_ws.sheet_state = "hidden"
    list_col_counter = {"value": 1}
    build_rounds_sheet(wb, list_ws, list_col_counter)
    build_hole_stats_sheet(wb, list_ws, list_col_counter)
    wb.save(TEMPLATE_PATH)
    print(f"Template created at {TEMPLATE_PATH}")


if __name__ == "__main__":
    main()
